import openpyxl
from os import path
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


# Abrir el archivo de excel
Nombre_archivo = 'Parametros caso NI 1.5.xlsm'
ruta = path.abspath(Nombre_archivo)
excel_document = openpyxl.load_workbook(ruta, keep_vba=True)
sheet = excel_document.get_sheet_by_name('Parametros de entrada')

# ---------------------------------------------------------------------------------------------------------

# conjuntos

fabricas = {1, 2, 3, 4}
bodegas = {1, 2}
productos = {1, 2, 3, 4}
clientes = {1, 2, 3, 4, 5, 6, 7, 8}
periodo = {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12}
periodo_mas_cero = {0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12}
periodo_menos_uno = {2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12}

# Extraccion de parametros

print "Cargando parametros"

# costos de ordenar a las fabricas
CO = {}
fila = 3
for j in bodegas:
    for f in fabricas:
        for i in productos:
            fila += 1
            CO[(i, f, j)] = float(sheet.cell(row=fila, column=4).value)


# costo de fabricar
CF = {}
fila = 3
for i in productos:
    for f in fabricas:
        fila += 1
        CF[(f, i)] = float(sheet.cell(row=fila, column=8).value)

# capacidad productiva
LO = {}
fila = 3
for i in productos:
    for f in fabricas:
        fila += 1
        LO[(f, i)] = sheet.cell(row=fila, column=12).value

# costo de transporte F-B
CT = {}
fila = 3
for j in bodegas:
    for i in productos:
        for f in fabricas:
            fila += 1
            CT[(f, i, j)] = float(sheet.cell(row=fila, column=17).value)

# costo de despacho B-C
CD = {}
fila = 3
for j in bodegas:
    for c in clientes:
        for i in productos:
            fila += 1
            CD[(i, c, j)] = float(sheet.cell(row=fila, column=22).value)


# demanda
D = {}
fila = 3
for c in clientes:
    for t in periodo:
        for i in productos:
            fila += 1
            D[(i, c, t)] = float(sheet.cell(row=fila, column=27).value)

# costo de almacenaje
CA = {}
fila = 3
for j in bodegas:
    for i in productos:
        fila += 1
        CA[(i, j)] = float(sheet.cell(row=fila, column=31).value)

# Inventario inicial
Inv = {}
fila = 3
for j in bodegas:
    for i in productos:
        fila += 1
        Inv[(i, j)] = sheet.cell(row=fila, column=35).value

# Capacidad de los camiones

CC = {}
fila = 3
for i in productos:
    fila += 1
    CC[i] = float(sheet.cell(row=fila, column=38).value)

# Costo fijo camion

CH = float(sheet.cell(row=2, column=41).value)

# gran M
M = 1000000000000


# ------------------------------------------------------------------------------------------------------------------
# modelo

import pyscipopt
from pyscipopt import Model, quicksum
m = pyscipopt.Model()

print "Resolviendo modelo"

# Variables

# punto de reorden producto i bodega j
x = {}
for i in productos:
    for j in bodegas:
        x[i, j] = m.addVar(vtype="C", lb=0, name="x(%s, %s)" % (i, j))

# up to level producto i bodega j
y = {}
for i in productos:
    for j in bodegas:
        y[i, j] = m.addVar(vtype="C", lb=0, name="y(%s, %s)" % (i, j))

# variable de estado de inventario producto i bodega j en periodo t
z = {}
for i in productos:
    for j in bodegas:
        for t in periodo_mas_cero:
            z[i, j, t] = m.addVar(vtype="C", name="z(%s, %s, %s)" % (i, j, t))

# variable de estado nivel de servicio producto i bodega j en periodo t
w = {}
for i in productos:
    for j in bodegas:
        for c in clientes:
            for t in periodo:
                w[i, j, c, t] = m.addVar(vtype="C", name="w(%s, %s, %s, %s)" % (i, j, c, t))

# 1 reposicion de producto i bodega j en periodo t. 0 en otro caso
q = {}
for i in productos:
    for j in bodegas:
        for t in periodo:
            q[i, j, t] = m.addVar(vtype="B", name="q(%s, %s, %s)" % (i, j, t))


# cuanto ordenar a fabrica f de producto i a bodega j en periodo t
r = {}
for f in fabricas:
    for i in productos:
        for j in bodegas:
            for t in periodo_mas_cero:
                r[f, i, j, t] = m.addVar(vtype="C", lb=0, name="r(%s, %s, %s, %s)" % (f, i, j, t))

# cuanto despachar de producto i de bodega j a cliente c en periodo t
s = {}
for i in productos:
    for j in bodegas:
        for c in clientes:
            for t in periodo:
                s[i, j, c, t] = m.addVar(vtype="C", lb=0, name="s(%s, %s, %s, %s)" % (i, j, c, t))

k = m.addVar("k", vtype="C", lb=None)

# Costo total de almacenamiento
costo_almacenamiento = m.addVar(vtype="C")

# costo total entre fabrica y bodega
costo_F_B = m.addVar(vtype="C")

# costo total entre bodega y cliente
costo_B_C = m.addVar(vtype="C")

# restricciones

# No puedo ordenar mas que la capacidad productiva de la fabrica.

for t in periodo:
    for i in productos:
        for f in fabricas:
            m.addCons(quicksum(r[f, i, j, t] for j in bodegas) <= LO[f, i])

# Lo que se ordena de producto debe sumar el up to level menos el nivel de inventario actual

for t in periodo:
    for i in productos:
        for j in bodegas:
            m.addCons(quicksum(r[f, i, j, t] for f in fabricas) == y[i, j] - z[i, j, t])

# No se puede generar ordenes si no es tiempo de reposicion (se utiliza una gran M)

for t in periodo:
    for i in productos:
        for j in bodegas:
            for f in fabricas:
                m.addCons(r[f, i, j, t] <= M*q[i, j, t])

# Activar el tiempo de reposicion (se utiliza una gran M)

for i in productos:
    for j in bodegas:
        for t in periodo:
            m.addCons(q[i, j, t] >= ((x[i, j]-z[i, j, t])/M))

# El up to level debe ser siempre superior o igual al punto de reorden

for i in productos:
    for j in bodegas:
        m.addCons(y[i, j] >= x[i, j])

# Las unidades despachadas no pueden superar el nivel de inventario actual

for i in productos:
    for j in bodegas:
        for t in periodo:
            m.addCons(quicksum(s[i, j, c, t] for c in clientes) <= z[i, j, t-1] + quicksum(r[f, i, j, t] for f in fabricas))

# Conservacion de flujo en las bodegas

# restriccion de borde de para inventario
for i in productos:
    for j in bodegas:
        m.addCons(z[i, j, 0] == Inv[i, j])

# restriccion de borde de para ordenes
for f in fabricas:
    for i in productos:
        for j in bodegas:
            m.addCons(r[f, i, j, 0] == 0)


for i in productos:
    for j in bodegas:
        for t in periodo:
            m.addCons(quicksum(r[f, i, j, t-1] for f in fabricas) + z[i, j, t-1] == z[i, j, t] + quicksum(s[i, j, c, t] for c in clientes))

# Armar variable de estado de nivel de servicio

for i in productos:
    for j in bodegas:
        for t in periodo:
            for c in clientes:
                m.addCons(w[i, j, c, t] == (z[i, j, t]/(quicksum(D[i, c, t] for c in clientes) + 1)))

# Satisfacer nivel de servicio

for i in productos:
    for j in bodegas:
        for t in periodo_menos_uno:
            for c in clientes:
                m.addCons(w[i, j, c, t] >= 0.95)

# satisfacer demanda de los clientes

for i in productos:
    for t in periodo:
        for c in clientes:
            m.addCons(quicksum(s[i, j, c, t] for j in bodegas) == D[i, c, t])


#funcion objetivo

m.addCons((quicksum((CO[i, f, j] + CF[f, i] + ((CT[f, i, j] + CH)/CC[i]))*r[f, i, j, t] for t in periodo for i in productos for j in bodegas for f in fabricas)
        + quicksum(CA[i, j]*z[i, j, t] for t in periodo for j in bodegas for i in productos)
        + quicksum(((CD[i, c, j] + CH)/CC[i])*s[i, j, c, t] for t in periodo for c in clientes for j in bodegas for i in productos)) == k)

m.addCons(quicksum(CA[i, j]*z[i, j, t] for t in periodo for j in bodegas for i in productos) == costo_almacenamiento)
m.addCons(quicksum((CO[i, f, j] + CF[f, i] + ((CT[f, i, j] + CH)/CC[i]))*r[f, i, j, t] for t in periodo for i in productos for j in bodegas for f in fabricas) == costo_F_B)
m.addCons(quicksum(((CD[i, c, j] + CH)/CC[i])*s[i, j, c, t] for t in periodo for c in clientes for j in bodegas for i in productos) == costo_B_C)


m.setObjective(k, "minimize")
m.optimize()
m.printBestSol()
print("El optimo es: ", m.getVal(k))

for i in productos:
    for j in bodegas:
        print (m.getVal(x[i, j]))

print ("El costo de almacenamiento es: ", m.getVal(costo_almacenamiento))
print ("El costo de Fabrica y Bodega es: ", m.getVal(costo_F_B))
print ("El costo de Bodega y Cliente es: ", m.getVal(costo_B_C))

for i in productos:
    for c in clientes:
        for t in periodo:
            print "reposicion producto "+str(i) + ", cliente " + str(c) + ", periodo " + str(t) + ": " + "bosdega 1 " + str(m.getVal(s[i, 1, c, t])) + " bodega 2 " + str(m.getVal(s[i, 2, c, t])) + " y la demanda " + str(D[i, c, t])

for t in periodo:
    print "reposicion producto 1, cliente 1, periodo " + str(t) + ": " + "bosdega 1 " + str(
        m.getVal(s[1, 1, 1, t])) + " bodega 2 " + str(m.getVal(s[1, 2, 1, t])) + " y la demanda " + str(D[1, 1, t])



# ----------------------------exportar resultados----------------------------------------------------------------------

# crear diccionario de resultados
resultados_y = {}
for i in productos:
    for j in bodegas:
        resultados_y[i, j] = m.getVal(y[i, j])

# Activate worksheet to write dataframe
active = excel_document['Resultados_y']

# Convertir diccionario a dataframe
df = pd.DataFrame(resultados_y.values(), index=pd.MultiIndex.from_tuples(resultados_y.keys())).unstack(1)

# Write dataframe to active worksheet
for i in dataframe_to_rows(df):
    active.append(i)

# ----------------------realizamos lo mismo para el resto de las variables--------------------------------------

# reposiciones bodega 1
Resultados_r1 = {}
for f in fabricas:
    for i in productos:
        valor = 0
        for t in periodo:
            valor = valor + m.getVal(r[f, i, 1, t])
        Resultados_r1[f, i] = valor

active = excel_document['Resultados_r1']
df = pd.DataFrame(Resultados_r1.values(), index=pd.MultiIndex.from_tuples(Resultados_r1.keys())).unstack(1)
for i in dataframe_to_rows(df):
    active.append(i)

# reposiciones bodega 2
Resultados_r2 = {}
for f in fabricas:
    for i in productos:
        valor = 0
        for t in periodo:
            valor = valor + m.getVal(r[f, i, 2, t])
        Resultados_r2[f, i] = valor

active = excel_document['Resultados_r2']
df = pd.DataFrame(Resultados_r2.values(), index=pd.MultiIndex.from_tuples(Resultados_r2.keys())).unstack(1)
for i in dataframe_to_rows(df):
    active.append(i)

# despacho a clientes

# despachos de la bodega 1
Resultados_s1 = {}
for i in productos:
    for c in clientes:
        valor = 0
        for t in periodo:
            valor = valor + m.getVal(s[i, 1, c, t])
        Resultados_s1[i, c] = valor

active = excel_document['Resultados_s1']
df = pd.DataFrame(Resultados_s1.values(), index=pd.MultiIndex.from_tuples(Resultados_s1.keys())).unstack(1)
for i in dataframe_to_rows(df):
    active.append(i)

# despachos de la bodega 2
Resultados_s2 = {}
for i in productos:
    for c in clientes:
        valor = 0
        for t in periodo:
            valor = valor + m.getVal(s[i, 2, c, t])
        Resultados_s2[i, c] = valor

active = excel_document['Resultados_s2']
df = pd.DataFrame(Resultados_s2.values(), index=pd.MultiIndex.from_tuples(Resultados_s2.keys())).unstack(1)
for i in dataframe_to_rows(df):
    active.append(i)

#costo total
Resultados_k = {}
Resultados_k[1, 1] = m.getVal(k)
Resultados_k[2, 1] = m.getVal(costo_almacenamiento)
Resultados_k[3, 1] = m.getVal(costo_F_B)
Resultados_k[4, 1] = m.getVal(costo_B_C)

active = excel_document['Resultados_k']
df = pd.DataFrame(Resultados_k.values(), index=pd.MultiIndex.from_tuples(Resultados_k.keys())).unstack(1)
for i in dataframe_to_rows(df):
    active.append(i)


# Save workbook to write
excel_document.save(ruta)

for t in periodo:
    for c in clientes:
        for i in productos:
            print "demanda producto " + str(i) + ", cliente " + str(c) + ", periodo " + str(t) + " : "+ str(D[i, c, t])

